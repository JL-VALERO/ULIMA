import openpyxl, os

base = r'C:\Users\51950\Documents\ULIMA\2026_1\Econometria 2\Trabajo_metria_2_2026_1'
folders = ['Variables macro chile', 'Data']

for folder in folders:
    path = os.path.join(base, folder)
    print(f'\n{"="*60}')
    print(f'CARPETA: {folder}')
    print(f'{"="*60}')
    for f in sorted(os.listdir(path)):
        if not f.endswith('.xlsx'):
            continue
        fpath = os.path.join(path, f)
        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        for sh in wb.sheetnames:
            ws = wb[sh]
            headers = [c.value for c in next(ws.iter_rows(max_row=1))]
            nrows = ws.max_row

            # Primera y ultima fila de datos
            all_rows = list(ws.iter_rows(values_only=True))
            first_data = all_rows[1] if len(all_rows) > 1 else None
            last_data  = all_rows[-1] if all_rows else None

            print(f'\n  Archivo : {f}')
            print(f'  Hoja    : {sh}  ({nrows} filas)')
            print(f'  Headers : {headers}')
            print(f'  1ra fila: {first_data}')
            print(f'  Ult fila: {last_data}')
        wb.close()
