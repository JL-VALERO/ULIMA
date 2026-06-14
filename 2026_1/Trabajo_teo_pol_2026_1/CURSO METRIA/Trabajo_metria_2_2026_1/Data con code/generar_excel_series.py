"""
Script: Descarga series macroeconómicas de Chile (mensual, 2000-presente)
Variables: Brent, IMACEC, Riesgo País (EMBI), TPM, M1, IPC
Autor generado: Claude Code
"""

import requests
import pandas as pd
from io import BytesIO, StringIO
import warnings
import sys

warnings.filterwarnings("ignore")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "es-CL,es;q=0.9,en;q=0.8",
}
BOLETIN = "https://si3.bcentral.cl/estadisticas/Principal1/enlaces/Informes/BOLETIN/listado/"
START = "2000-01-01"
END   = "2026-04-30"

series_log = []   # para el sheet de metadatos

# ─────────────────────────────────────────────────────────────────────
def log(var, tipo, unidad, freq, fuente, rango=""):
    series_log.append({
        "Variable": var,
        "Tipo (Índice/Variación)": tipo,
        "Unidad": unidad,
        "Frecuencia": freq,
        "Fuente": fuente,
        "Cobertura temporal": rango,
    })

# ─────────────────────────────────────────────────────────────────────
# 1. BRENT – EIA mensual (USD/barril, NIVEL)
# ─────────────────────────────────────────────────────────────────────
def get_brent():
    print("1. Brent (EIA)...")
    try:
        r = requests.get(
            "https://www.eia.gov/dnav/pet/xls/PET_PRI_SPT_S1_M.xls",
            headers=HEADERS, timeout=60
        )
        r.raise_for_status()
        df = pd.read_excel(BytesIO(r.content), sheet_name="Data 1", skiprows=2)
        # Buscar la columna de Brent (Europe)
        brent_col = next(
            (c for c in df.columns if "brent" in str(c).lower() or "europe" in str(c).lower()),
            df.columns[1]
        )
        df = df[["Date", brent_col]].copy()
        df.columns = ["fecha", "brent_usd_bbl"]
        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
        df = df.dropna(subset=["fecha"])
        df = df[(df["fecha"] >= START) & (df["fecha"] <= END)]
        df["fecha"] = df["fecha"].dt.to_period("M").dt.to_timestamp()
        df = df.sort_values("fecha").reset_index(drop=True)
        rango = f"{df['fecha'].min().strftime('%Y-%m')} – {df['fecha'].max().strftime('%Y-%m')}"
        log("Precio Brent", "NIVEL (Índice de Precio)", "USD por barril",
            "Mensual", "EIA (US Energy Information Administration)", rango)
        print(f"   OK – {len(df)} obs ({rango})")
        return df
    except Exception as e:
        print(f"   ERROR: {e}")
        log("Precio Brent", "NIVEL", "USD por barril", "Mensual", "EIA", "ERROR")
        return pd.DataFrame(columns=["fecha", "brent_usd_bbl"])


# ─────────────────────────────────────────────────────────────────────
# 2a. IMACEC – Variación anual (BC001_g.xls)
# ─────────────────────────────────────────────────────────────────────
def parse_boletin_variacion(fname, col_idx, col_name, skip_rows=8):
    """Lee un Excel del Boletín y extrae una columna de variación mensual."""
    r = requests.get(BOLETIN + fname, headers=HEADERS, timeout=60)
    r.raise_for_status()
    df_raw = pd.read_excel(BytesIO(r.content), sheet_name=0, header=None)
    df = pd.DataFrame({
        "fecha": df_raw.iloc[skip_rows:, 0].values,
        col_name: df_raw.iloc[skip_rows:, col_idx].values,
    })
    df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
    df = df.dropna(subset=["fecha"])
    df = df[(df["fecha"] >= START) & (df["fecha"] <= END)]
    df["fecha"] = df["fecha"].dt.to_period("M").dt.to_timestamp()
    return df.sort_values("fecha").reset_index(drop=True)

def get_imacec():
    print("2. IMACEC (BCCh Boletín)...")
    result = {}
    # a) Variación anual (% respecto al mismo mes del año anterior)
    try:
        df_a = parse_boletin_variacion("BC001_g.xls", col_idx=9, col_name="imacec_var_anual_pct")
        rango = f"{df_a['fecha'].min().strftime('%Y-%m')} – {df_a['fecha'].max().strftime('%Y-%m')}"
        log("IMACEC", "VARIACIÓN (% respecto mismo mes año anterior)",
            "Porcentaje (%)", "Mensual",
            "Banco Central de Chile (BCCh) – Boletín Estadístico BC001_g.xls", rango)
        print(f"   OK variación anual – {len(df_a)} obs ({rango})")
        result["var_anual"] = df_a
    except Exception as e:
        print(f"   ERROR var anual: {e}")
        result["var_anual"] = pd.DataFrame(columns=["fecha", "imacec_var_anual_pct"])

    # b) Variación mensual desestacionalizada (% respecto mes anterior)
    try:
        df_m = parse_boletin_variacion("BC001_2.xls", col_idx=9, col_name="imacec_var_mensual_desest_pct")
        rango = f"{df_m['fecha'].min().strftime('%Y-%m')} – {df_m['fecha'].max().strftime('%Y-%m')}"
        log("IMACEC desestacionalizado", "VARIACIÓN (% respecto mes anterior, desest.)",
            "Porcentaje (%)", "Mensual",
            "Banco Central de Chile (BCCh) – Boletín Estadístico BC001_2.xls", rango)
        print(f"   OK variación mensual desest. – {len(df_m)} obs ({rango})")
        result["var_mensual"] = df_m
    except Exception as e:
        print(f"   ERROR var mensual: {e}")
        result["var_mensual"] = pd.DataFrame(columns=["fecha", "imacec_var_mensual_desest_pct"])

    return result


# ─────────────────────────────────────────────────────────────────────
# 3. EMBI Chile – BCCh BDE (requiere credenciales) / BCRP 2017+
# ─────────────────────────────────────────────────────────────────────
def get_embi():
    print("3. EMBI Chile (riesgo país)...")
    # Intentar BCCh Siete REST (sin credenciales)
    bcch_series_ids = [
        "F074.EMBI.EMB.CHI.M",
        "F074.EMBI.EMB.CHI.D.N",
        "F047.EMB.CHI.M",
    ]
    for sid in bcch_series_ids:
        try:
            url = (
                f"https://si3.bcentral.cl/SieteRestWS/SieteRestWS.ashx?"
                f"user=&pass=&function=GetSeries&timeseries={sid}"
                f"&firstdate=2000-01-01&lastdate=2026-04-01&freq=M&format=JSON"
            )
            r = requests.get(url, headers=HEADERS, timeout=20)
            if r.status_code == 200 and "Series" in r.text:
                data = r.json()
                print(f"   BCCh API ok con serie {sid}:", str(data)[:200])
                break
        except Exception:
            pass

    # BCRP: PN01132XM – datos disponibles Ene-2017 a Dic-2025
    try:
        url_bcrp = (
            "https://estadisticas.bcrp.gob.pe/estadisticas/series/api/"
            "PN01132XM/xls/mensual/Ene-2017/Dic-2025"
        )
        r = requests.get(url_bcrp, headers={**HEADERS, "Referer": "https://estadisticas.bcrp.gob.pe/"}, timeout=30)
        if r.status_code == 200 and len(r.content) > 5000:
            xls = pd.ExcelFile(BytesIO(r.content))
            # Los datos deberían estar en la hoja 'Api'
            df_raw = pd.read_excel(BytesIO(r.content), sheet_name="Api", header=0)
            print("   BCRP Api sheet:", df_raw.shape, df_raw.columns.tolist())
            print(df_raw.head(5).to_string())
    except Exception as e:
        print(f"   BCRP error: {e}")

    # Crear DataFrame vacío con nota
    print("   NOTA: EMBI Chile no disponible sin credenciales BCCh BDE.")
    print("         Ver instrucciones en sheet 'Metadatos' del Excel.")
    log("EMBI Chile (Riesgo País)", "NIVEL (spread en puntos base)",
        "Puntos base (pb)", "Mensual",
        "BCCh BDE – requiere registro en https://si3.bcentral.cl/SieteRestWS/",
        "N/D – requiere credenciales")
    return pd.DataFrame(columns=["fecha", "embi_chile_pb"])


# ─────────────────────────────────────────────────────────────────────
# 4. TPM – Tasa de Política Monetaria (BB032_2_g.xls, diaria → promedio mensual)
# ─────────────────────────────────────────────────────────────────────
def get_tpm():
    print("4. TPM (BCCh Boletín)...")
    try:
        r = requests.get(BOLETIN + "BB032_2_g.xls", headers=HEADERS, timeout=60)
        r.raise_for_status()
        df_raw = pd.read_excel(BytesIO(r.content), sheet_name=0, header=None)
        df = pd.DataFrame({
            "fecha": df_raw.iloc[8:, 0].values,
            "tpm_pct": df_raw.iloc[8:, 1].values,
        })
        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
        df = df.dropna(subset=["fecha"])
        df["tpm_pct"] = pd.to_numeric(df["tpm_pct"], errors="coerce")
        # Forward fill: la tasa vigente se mantiene hasta el próximo cambio
        df = df.sort_values("fecha")
        df["tpm_pct"] = df["tpm_pct"].ffill()
        df = df[(df["fecha"] >= START) & (df["fecha"] <= END)]
        # Promedio mensual (toma el valor de fin de mes, que es el vigente ese mes)
        df["mes"] = df["fecha"].dt.to_period("M").dt.to_timestamp()
        df_m = (
            df.groupby("mes")["tpm_pct"]
            .last()          # último valor registrado del mes = tasa vigente
            .reset_index()
            .rename(columns={"mes": "fecha"})
        )
        df_m = df_m.sort_values("fecha").reset_index(drop=True)
        rango = f"{df_m['fecha'].min().strftime('%Y-%m')} – {df_m['fecha'].max().strftime('%Y-%m')}"
        log("TPM (Tasa de Política Monetaria)", "NIVEL (tasa nominal anual)",
            "Porcentaje anual (%)", "Mensual (promedio de datos diarios)",
            "BCCh – Boletín Estadístico BB032_2_g.xls", rango)
        print(f"   OK – {len(df_m)} obs ({rango})")
        return df_m
    except Exception as e:
        print(f"   ERROR: {e}")
        log("TPM", "NIVEL", "%", "Mensual", "BCCh Boletín", "ERROR")
        return pd.DataFrame(columns=["fecha", "tpm_pct"])


# ─────────────────────────────────────────────────────────────────────
# 5. M1 – Agregados Monetarios (BB003.xls, columna 5 = M1, nivel)
# ─────────────────────────────────────────────────────────────────────
def get_m1():
    print("5. M1 (BCCh Boletín)...")
    try:
        r = requests.get(BOLETIN + "BB003.xls", headers=HEADERS, timeout=60)
        r.raise_for_status()
        df_raw = pd.read_excel(BytesIO(r.content), sheet_name=0, header=None)
        df = pd.DataFrame({
            "fecha": df_raw.iloc[8:, 0].values,
            "m1_mm_clp": df_raw.iloc[8:, 5].values,   # columna 5 = M1
        })
        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
        df = df.dropna(subset=["fecha"])
        df["m1_mm_clp"] = pd.to_numeric(df["m1_mm_clp"], errors="coerce")
        df = df[(df["fecha"] >= START) & (df["fecha"] <= END)]
        df["fecha"] = df["fecha"].dt.to_period("M").dt.to_timestamp()
        df = df.sort_values("fecha").reset_index(drop=True)
        rango = f"{df['fecha'].min().strftime('%Y-%m')} – {df['fecha'].max().strftime('%Y-%m')}"
        log("M1 (Agregado Monetario)", "NIVEL",
            "Miles de millones de pesos CLP (promedio mensual)",
            "Mensual",
            "BCCh – Boletín Estadístico BB003.xls", rango)
        print(f"   OK – {len(df)} obs ({rango})")
        return df
    except Exception as e:
        print(f"   ERROR: {e}")
        log("M1", "NIVEL", "Miles de MM CLP", "Mensual", "BCCh Boletín", "ERROR")
        return pd.DataFrame(columns=["fecha", "m1_mm_clp"])


# ─────────────────────────────────────────────────────────────────────
# 6. IPC – FRED (índice 2015=100) + extensión con variación BCCh b02_t.xls
# ─────────────────────────────────────────────────────────────────────
def get_ipc():
    print("6. IPC (FRED + extensión BCCh Boletín)...")
    df_ipc_var_mm = pd.DataFrame(columns=["fecha", "ipc_var_mensual_pct"])

    # 6a. Variación mensual IPC desde b02_t.xls (col 15 = IPC Base m/m)
    try:
        r = requests.get(BOLETIN + "b02_t.xls", headers=HEADERS, timeout=60)
        r.raise_for_status()
        df_raw = pd.read_excel(BytesIO(r.content), sheet_name=0, header=None)
        df_var = pd.DataFrame({
            "fecha": df_raw.iloc[8:, 0].values,
            "ipc_var_mensual_pct": df_raw.iloc[8:, 15].values,  # col 15 = IPC Base m/m
        })
        df_var["fecha"] = pd.to_datetime(df_var["fecha"], errors="coerce")
        df_var = df_var.dropna(subset=["fecha"])
        df_var["ipc_var_mensual_pct"] = pd.to_numeric(df_var["ipc_var_mensual_pct"], errors="coerce")
        df_var["fecha"] = df_var["fecha"].dt.to_period("M").dt.to_timestamp()
        df_var = df_var.sort_values("fecha").reset_index(drop=True)
        df_ipc_var_mm = df_var
        print(f"   Var. mensual IPC – {len(df_var)} obs "
              f"({df_var['fecha'].min().strftime('%Y-%m')} – "
              f"{df_var['fecha'].max().strftime('%Y-%m')})")
    except Exception as e:
        print(f"   Var. mensual IPC error: {e}")

    # 6b. Índice FRED/OECD (base 2015=100) – hasta dic 2023 aprox.
    df_fred = pd.DataFrame(columns=["fecha", "ipc_indice_2015_100"])
    try:
        import pandas_datareader as pdr
        df_fred = pdr.get_data_fred("CHLCPIALLMINMEI", start="2000-01-01", end="2026-04-30")
        df_fred = df_fred.reset_index()
        df_fred.columns = ["fecha", "ipc_indice_2015_100"]
        df_fred["fecha"] = pd.to_datetime(df_fred["fecha"])
        df_fred["fecha"] = df_fred["fecha"].dt.to_period("M").dt.to_timestamp()
        df_fred = df_fred.sort_values("fecha").reset_index(drop=True)
        print(f"   Índice FRED – {len(df_fred)} obs "
              f"({df_fred['fecha'].min().strftime('%Y-%m')} – "
              f"{df_fred['fecha'].max().strftime('%Y-%m')})")
    except Exception as e:
        print(f"   FRED error: {e}")

    # 6c. Extender índice FRED usando variación mensual BCCh para 2024+
    if len(df_fred) > 0 and len(df_ipc_var_mm) > 0:
        last_fred_date = df_fred["fecha"].max()
        # Variaciones mensuales después del último dato FRED
        df_ext = df_ipc_var_mm[df_ipc_var_mm["fecha"] > last_fred_date].copy()
        if len(df_ext) > 0:
            last_idx = float(df_fred.loc[df_fred["fecha"] == last_fred_date, "ipc_indice_2015_100"].values[0])
            extension_rows = []
            for _, row in df_ext.iterrows():
                if pd.notna(row["ipc_var_mensual_pct"]):
                    last_idx = last_idx * (1 + row["ipc_var_mensual_pct"] / 100)
                    extension_rows.append({"fecha": row["fecha"], "ipc_indice_2015_100": round(last_idx, 4)})
            if extension_rows:
                df_ext2 = pd.DataFrame(extension_rows)
                df_fred = pd.concat([df_fred, df_ext2], ignore_index=True)
                df_fred = df_fred.sort_values("fecha").reset_index(drop=True)
                print(f"   Índice extendido – total {len(df_fred)} obs, "
                      f"hasta {df_fred['fecha'].max().strftime('%Y-%m')}")

    df_fred = df_fred[(df_fred["fecha"] >= START) & (df_fred["fecha"] <= END)]
    df_fred = df_fred.sort_values("fecha").reset_index(drop=True)

    if len(df_fred) > 0:
        rango = f"{df_fred['fecha'].min().strftime('%Y-%m')} – {df_fred['fecha'].max().strftime('%Y-%m')}"
        log("IPC (Índice de Precios al Consumidor)", "ÍNDICE (base 2015=100, escala OCDE)",
            "Índice 2015=100",
            "Mensual",
            "FRED/OCDE (CHLCPIALLMINMEI) hasta 2023; extendido con Var.% BCCh b02_t.xls para 2024+",
            rango)
        # También agregar la variación mensual al log
        if len(df_ipc_var_mm) > 0:
            rango_v = (f"{df_ipc_var_mm['fecha'].min().strftime('%Y-%m')} – "
                       f"{df_ipc_var_mm['fecha'].max().strftime('%Y-%m')}")
            log("IPC Variación Mensual", "VARIACIÓN (% respecto al mes anterior)",
                "Porcentaje (%)", "Mensual",
                "BCCh – Boletín Estadístico b02_t.xls (col. IPC Base m/m)", rango_v)
        print(f"   OK IPC final – {len(df_fred)} obs ({rango})")
    else:
        log("IPC", "ÍNDICE", "Índice", "Mensual", "FRED/OCDE – ERROR", "")

    # Merge índice + variación mensual
    result = df_fred
    if len(df_ipc_var_mm) > 0:
        df_ipc_var_mm = df_ipc_var_mm[(df_ipc_var_mm["fecha"] >= START) & (df_ipc_var_mm["fecha"] <= END)]
        result = result.merge(df_ipc_var_mm, on="fecha", how="outer")
        result = result.sort_values("fecha").reset_index(drop=True)
    return result


# ─────────────────────────────────────────────────────────────────────
# 7. EMBI Complementario – intento BCCh REST (público)
# ─────────────────────────────────────────────────────────────────────
def try_bcch_rest_embi():
    """Intenta descargar el EMBI desde la API pública del BCCh (sin credenciales).
    Retorna None si no funciona."""
    # Varios series IDs posibles para el EMBI Chile mensual
    candidates = [
        "F074.EMBI.EMB.CHI.M",
        "F047.EMP.EMB.CHI.M",
        "F074.EMBI.CHI.D.M",
    ]
    for sid in candidates:
        try:
            url = (
                "https://si3.bcentral.cl/SieteRestWS/SieteRestWS.ashx?"
                f"user=none&pass=none&function=GetSeries&timeseries={sid}"
                "&firstdate=2000-01-01&lastdate=2026-04-01&freq=M&format=JSON"
            )
            r = requests.get(url, headers=HEADERS, timeout=15)
            if r.status_code == 200 and '"Series"' in r.text:
                data = r.json()
                if data.get("Series"):
                    rows = data["Series"][0].get("Obs", [])
                    if rows:
                        df = pd.DataFrame(rows)
                        df = df.rename(columns={"indexDateString": "fecha", "value": "embi_chile_pb"})
                        df["fecha"] = pd.to_datetime(df["fecha"], errors="coerce")
                        df = df.dropna(subset=["fecha"])
                        df["embi_chile_pb"] = pd.to_numeric(df["embi_chile_pb"], errors="coerce")
                        df = df[["fecha", "embi_chile_pb"]].sort_values("fecha").reset_index(drop=True)
                        print(f"   BCCh REST OK con serie {sid}: {len(df)} obs")
                        return df
        except Exception:
            pass
    return None


# ─────────────────────────────────────────────────────────────────────
# MAIN – Combinar todo en un Excel
# ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("=" * 65)
    print("Descargando series macroeconómicas Chile (2000-2026)")
    print("=" * 65)

    df_brent     = get_brent()
    imacec_dict  = get_imacec()
    df_embi      = get_embi()
    df_tpm       = get_tpm()
    df_m1        = get_m1()
    df_ipc       = get_ipc()

    # Intentar EMBI desde BCCh REST (sin credenciales)
    embi_rest = try_bcch_rest_embi()
    if embi_rest is not None and len(embi_rest) > 0:
        df_embi = embi_rest
        # Actualizar log
        rango_embi = f"{df_embi['fecha'].min().strftime('%Y-%m')} – {df_embi['fecha'].max().strftime('%Y-%m')}"
        for row in series_log:
            if "EMBI" in row["Variable"]:
                row["Cobertura temporal"] = rango_embi
                row["Fuente"] = "BCCh BDE REST API"

    # ── Merge de todas las series sobre una base de fechas mensuales ──
    print("\nCombinando series...")
    all_fechas = set()
    for df_temp in [df_brent, imacec_dict.get("var_anual", pd.DataFrame()),
                    imacec_dict.get("var_mensual", pd.DataFrame()),
                    df_embi, df_tpm, df_m1, df_ipc]:
        if "fecha" in df_temp.columns:
            all_fechas.update(df_temp["fecha"].dropna().tolist())

    df_base = pd.DataFrame({"fecha": sorted(all_fechas)})
    df_base = df_base[(df_base["fecha"] >= START) & (df_base["fecha"] <= END)]

    for df_temp in [
        df_brent,
        imacec_dict.get("var_anual", pd.DataFrame()),
        imacec_dict.get("var_mensual", pd.DataFrame()),
        df_embi,
        df_tpm,
        df_m1,
        df_ipc,
    ]:
        if "fecha" in df_temp.columns and len(df_temp) > 0:
            df_base = df_base.merge(df_temp, on="fecha", how="left")

    df_base = df_base.sort_values("fecha").reset_index(drop=True)

    # ── Renombrar columnas para mayor claridad en Excel ──
    rename_map = {
        "brent_usd_bbl":                     "Brent (USD/barril) [NIVEL]",
        "imacec_var_anual_pct":              "IMACEC Var.Anual (% a/a) [VARIACIÓN]",
        "imacec_var_mensual_desest_pct":     "IMACEC Var.Mensual Desest. (% m/m) [VARIACIÓN]",
        "embi_chile_pb":                     "EMBI Chile (puntos base) [NIVEL]",
        "tpm_pct":                           "TPM (% anual) [NIVEL]",
        "m1_mm_clp":                         "M1 (miles de MM CLP) [NIVEL]",
        "ipc_indice_2015_100":               "IPC Índice OCDE (2015=100) [ÍNDICE]",
        "ipc_var_mensual_pct":               "IPC Var.Mensual (% m/m) [VARIACIÓN]",
    }
    df_base = df_base.rename(columns=rename_map)

    # ── Construir hoja de metadatos ──
    df_meta = pd.DataFrame(series_log)
    df_meta_extra = pd.DataFrame([
        {
            "Variable": "─── NOTAS ───",
            "Tipo (Índice/Variación)": "",
            "Unidad": "",
            "Frecuencia": "",
            "Fuente": "",
            "Cobertura temporal": "",
        },
        {
            "Variable": "IPC",
            "Tipo (Índice/Variación)": "ÍNDICE",
            "Unidad": "Base 2015=100 (escala OCDE). Para convertir a base INE-Chile 2023=100, dividir por el valor de 2023 y multiplicar por 100.",
            "Frecuencia": "",
            "Fuente": "OCDE / FRED",
            "Cobertura temporal": "Hasta diciembre 2023 aprox. Para datos más recientes: https://www.ine.gob.cl/estadisticas/economia/indices-de-precio-e-inflacion/indice-de-precios-al-consumidor",
        },
        {
            "Variable": "EMBI Chile",
            "Tipo (Índice/Variación)": "NIVEL (spread)",
            "Unidad": "Para obtener la serie completa desde 2000: registrarse gratis en https://si3.bcentral.cl/SieteRestWS/index.html y descargar la serie del EMBI Chile.",
            "Frecuencia": "",
            "Fuente": "BCCh BDE",
            "Cobertura temporal": "Alternativa parcial (2017+): https://estadisticas.bcrp.gob.pe → EMBI → Chile → serie PN01132XM",
        },
        {
            "Variable": "IMACEC",
            "Tipo (Índice/Variación)": "VARIACIÓN",
            "Unidad": "El Índice IMACEC (base 2013=100) mensual está disponible en BCCh BDE (requiere registro).",
            "Frecuencia": "",
            "Fuente": "BCCh Boletín",
            "Cobertura temporal": "Para el índice mensual: https://si3.bcentral.cl → IMACEC encadenado",
        },
    ])
    df_meta = pd.concat([df_meta, df_meta_extra], ignore_index=True)

    # ── Guardar Excel ──
    output_path = "Series_Macroeconomicas_Chile_2000_2026.xlsx"
    print(f"\nGuardando: {output_path} ...")
    with pd.ExcelWriter(output_path, engine="openpyxl", datetime_format="YYYY-MM") as writer:
        df_base.to_excel(writer, sheet_name="Series_Tiempo", index=False)
        df_meta.to_excel(writer, sheet_name="Metadatos_Fuentes", index=False)

        # Ajustar anchos de columna
        from openpyxl.utils import get_column_letter
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        import openpyxl

        wb = writer.book

        # ── Formato hoja Series_Tiempo ──
        ws_series = wb["Series_Tiempo"]
        # Header row styling
        header_fill = PatternFill("solid", fgColor="1F4E79")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for cell in ws_series[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Ajustar anchos
        ws_series.column_dimensions["A"].width = 12  # fecha
        for col_idx in range(2, ws_series.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws_series.column_dimensions[col_letter].width = 28

        # Formato alternado de filas
        fill_alt = PatternFill("solid", fgColor="D9E1F2")
        for row_idx, row in enumerate(ws_series.iter_rows(min_row=2), start=2):
            if row_idx % 2 == 0:
                for cell in row:
                    cell.fill = fill_alt
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # Freeze header
        ws_series.freeze_panes = "A2"

        # ── Formato hoja Metadatos ──
        ws_meta = wb["Metadatos_Fuentes"]
        header_fill2 = PatternFill("solid", fgColor="375623")
        for cell in ws_meta[1]:
            cell.fill = header_fill2
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        for col_idx in range(1, ws_meta.max_column + 1):
            col_letter = get_column_letter(col_idx)
            ws_meta.column_dimensions[col_letter].width = 35
        for row in ws_meta.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    print(f"\n{'=' * 65}")
    print(f"Excel generado exitosamente: {output_path}")
    print(f"  Filas de datos: {len(df_base)}")
    print(f"  Columnas: {list(df_base.columns)}")
    print(f"\nResumen de cobertura:")
    for entry in series_log:
        print(f"  {entry['Variable'][:45]:<45} | {entry['Cobertura temporal']}")
    print(f"\nSeries incluidas:")
    print(f"  NIVEL   (valor absoluto): Brent, TPM, M1, IPC, EMBI (si disponible)")
    print(f"  VARIACIÓN (cambio %):     IMACEC Var. Anual, IMACEC Var. Mensual Desest.")
    print(f"{'=' * 65}")
