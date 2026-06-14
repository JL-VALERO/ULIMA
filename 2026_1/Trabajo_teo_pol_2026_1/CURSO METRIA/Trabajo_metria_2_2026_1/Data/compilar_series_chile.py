"""
Compila 6 series mensuales de Chile: Ene 2004 - Dic 2025
1. IPC          - Indice nivel (base 2023=100)           -> IPC Chile indice bases 2023.xlsx
2. PBI/IMACEC   - IMACEC desestacionalizado (base 2018)  -> IMACEC Chile.xlsx
3. Tasa Interm. - TPM (% anual)                          -> TIR Chile.xlsx
4. Tipo Cambio  - CLP/USD promedio mensual               -> DATA_JORGE.xlsx col 'tc'
5. Reservas     - Reservas Int. Netas (mill. USD)        -> BCCh Reservas_internacionales.xls
6. Credito      - Colocaciones MN total (mill. CLP)      -> BCCh BB052.xls (desde 1990)
"""

import os, sys, requests, warnings
import pandas as pd
import numpy as np
import openpyxl, xlrd
from io import BytesIO

warnings.filterwarnings("ignore")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")

BASE    = os.path.dirname(os.path.abspath(__file__))
BASE_UP = os.path.join(BASE, "..")
VMC     = os.path.join(BASE_UP, "Variables macro chile")
START   = pd.Timestamp("2004-01-01")
END     = pd.Timestamp("2025-12-01")
HDR     = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

# Carpeta destino: Regresion_Var_IPC/datos (usuario ya movio el Excel ahi)
DEST_DIR = os.path.normpath(os.path.join(
    BASE_UP, "..", "..", "..", "Regresion_Var_IPC", "datos"))
if not os.path.isdir(DEST_DIR):
    # Fallback: misma carpeta Data
    DEST_DIR = BASE

OUT_FILE = os.path.join(DEST_DIR, "Series_Chile_2004_2025.xlsx")

def filtrar(df, col="fecha"):
    df[col] = pd.to_datetime(df[col], errors="coerce")
    df = df[(df[col] >= START) & (df[col] <= END)].dropna(subset=[col])
    df[col] = df[col].dt.to_period("M").dt.to_timestamp()
    return df.sort_values(col).reset_index(drop=True)

def leer_xlsx_col(filepath, sheet, col_fecha=0, col_valor=1, skip=1):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    rows = list(wb[sheet].iter_rows(values_only=True))
    wb.close()
    data = [(r[col_fecha], r[col_valor]) for r in rows[skip:]
            if r[col_fecha] is not None and r[col_valor] is not None]
    return pd.DataFrame(data, columns=["fecha", "valor"])

print("=" * 62)
print("COMPILACION DE SERIES MACROECONOMICAS DE CHILE 2004-2025")
print("=" * 62)

# ── 1. IPC ────────────────────────────────────────────────────
print("\n[1] IPC (base 2023=100) - archivo local")
df_ipc = filtrar(leer_xlsx_col(
    os.path.join(VMC, "IPC Chile indice bases 2023 y otro.xlsx"), "Tabla1")
).rename(columns={"valor": "ipc_indice"})
print(f"    OK - {len(df_ipc)} obs | {df_ipc['ipc_indice'].min():.1f} - {df_ipc['ipc_indice'].max():.1f}")

# ── 2. IMACEC (PBI mensual) ───────────────────────────────────
print("\n[2] IMACEC desest. (base 2018=100) - archivo local")
df_imacec = filtrar(leer_xlsx_col(
    os.path.join(BASE, "IMACEC Chile.xlsx"), "Cuadro")
).rename(columns={"valor": "imacec_pbi"})
print(f"    OK - {len(df_imacec)} obs | {df_imacec['imacec_pbi'].min():.1f} - {df_imacec['imacec_pbi'].max():.1f}")

# ── 3. Tasa Interbancaria (TPM) ───────────────────────────────
print("\n[3] TPM (% anual) - archivo local")
df_tpm = filtrar(leer_xlsx_col(
    os.path.join(BASE, "TIR Chile.xlsx"), "Cuadro")
).rename(columns={"valor": "tasa_interbancaria_tpm"})
print(f"    OK - {len(df_tpm)} obs | {df_tpm['tasa_interbancaria_tpm'].min():.2f}% - {df_tpm['tasa_interbancaria_tpm'].max():.2f}%")

# ── 4. Tipo de Cambio CLP/USD ─────────────────────────────────
print("\n[4] TC CLP/USD - DATA_JORGE.xlsx")
wb = openpyxl.load_workbook(os.path.join(BASE, "DATA_JORGE.xlsx"),
                            read_only=True, data_only=True)
rows_j = list(wb["Sheet1"].iter_rows(values_only=True))
wb.close()
import datetime as _dt
data_tc = [(r[0], r[9]) for r in rows_j[3:]
           if isinstance(r[0], _dt.datetime) and r[9] is not None]
df_tc = filtrar(pd.DataFrame(data_tc, columns=["fecha", "tc_clp_usd"]))
print(f"    OK - {len(df_tc)} obs | {df_tc['tc_clp_usd'].min():.1f} - {df_tc['tc_clp_usd'].max():.1f} CLP/USD")

# ── 5. RESERVAS NETAS INTERNACIONALES ────────────────────────
print("\n[5] Reservas Int. Netas (mill. USD) - BCCh web")
url_res = "https://si3.bcentral.cl/estadisticas/Principal1/excel/EMF/RESERVAS/Excel/Reservas_internacionales.xls"
try:
    r = requests.get(url_res, headers=HDR, timeout=25)
    r.raise_for_status()
    wb_res = xlrd.open_workbook(file_contents=r.content)
    sh_res = wb_res.sheets()[0]
    # Estructura: fila 5 = cabeceras meses, filas 7+ = año + datos mensuales
    meses = sh_res.row_values(5)[1:]  # Enero, Febrero, ..., Diciembre
    records = []
    for row_idx in range(7, sh_res.nrows):
        vals = sh_res.row_values(row_idx)
        anio = vals[0]
        if not anio or not isinstance(anio, (int, float)) or anio < 2000:
            continue
        anio = int(anio)
        for col_idx, mes_nombre in enumerate(meses, start=1):
            if col_idx >= len(vals):
                break
            v = vals[col_idx]
            if v == '' or v is None:
                continue
            try:
                v = float(v)
            except (ValueError, TypeError):
                continue
            # Mapear nombre de mes a numero
            mes_map = {"Enero":1,"Febrero":2,"Marzo":3,"Abril":4,"Mayo":5,"Junio":6,
                       "Julio":7,"Agosto":8,"Septiembre":9,"Octubre":10,"Noviembre":11,"Diciembre":12}
            mes_num = mes_map.get(str(mes_nombre).strip(), None)
            if mes_num is None:
                continue
            records.append((pd.Timestamp(year=anio, month=mes_num, day=1), v))
    df_res = pd.DataFrame(records, columns=["fecha", "reservas_netas_musd"])
    df_res = filtrar(df_res)
    print(f"    OK - {len(df_res)} obs | {df_res['reservas_netas_musd'].min():.0f} - {df_res['reservas_netas_musd'].max():.0f} mill. USD")
except Exception as e:
    print(f"    ERROR: {e}")
    df_res = None

# ── 6. CREDITO PRIVADO (Colocaciones) ────────────────────────
print("\n[6] Credito Privado - Colocaciones MN total (mill. CLP) - BCCh BB052")
url_cred = "https://si3.bcentral.cl/estadisticas/Principal1/enlaces/Informes/BOLETIN/listado/BB052.xls"
try:
    r = requests.get(url_cred, headers=HDR, timeout=25)
    r.raise_for_status()
    wb_cred = xlrd.open_workbook(file_contents=r.content)
    sh_cred = wb_cred.sheets()[0]
    # Col 0 = fecha serial Excel, Col 8 = Colocaciones MN Total (mill. CLP)
    records_c = []
    for row_idx in range(8, sh_cred.nrows):
        serial = sh_cred.cell_value(row_idx, 0)
        val = sh_cred.cell_value(row_idx, 8)  # Col 8: MN Total
        if not serial or not val:
            continue
        try:
            dt = xlrd.xldate_as_datetime(serial, wb_cred.datemode)
            val = float(val)
            records_c.append((pd.Timestamp(year=dt.year, month=dt.month, day=1), val))
        except Exception:
            continue
    df_cred = pd.DataFrame(records_c, columns=["fecha", "credito_privado_mclp"])
    df_cred = filtrar(df_cred)
    n_miss = 264 - len(df_cred)
    print(f"    OK - {len(df_cred)} obs | {df_cred['credito_privado_mclp'].min():.0f} - {df_cred['credito_privado_mclp'].max():.0f} mill. CLP")
    if n_miss > 0:
        print(f"    Nota: {n_miss} meses finales sin publicar aun (BCCh lag ~2 meses)")
except Exception as e:
    print(f"    ERROR: {e}")
    df_cred = None

# ── CONSOLIDAR ────────────────────────────────────────────────
print("\n" + "=" * 62)
print("RESUMEN FINAL:")

series_list = [
    ("ipc_indice",            df_ipc,    "IPC nivel base 2023=100",           "Archivo local"),
    ("imacec_pbi",            df_imacec, "IMACEC desest. base 2018=100",      "Archivo local"),
    ("tasa_interbancaria_tpm",df_tpm,    "TPM (% anual)",                     "Archivo local"),
    ("tc_clp_usd",            df_tc,     "Tipo cambio CLP/USD prom. mensual", "DATA_JORGE.xlsx"),
    ("reservas_netas_musd",   df_res,    "Reservas Int. Netas (mill. USD)",   "BCCh Reservas_internacionales.xls"),
    ("credito_privado_mclp",  df_cred,   "Colocaciones MN total (mill. CLP)", "BCCh BB052.xls"),
]

maestro = pd.DataFrame({"fecha": pd.date_range("2004-01-01", "2025-12-01", freq="MS")})
maestro["periodo"] = maestro["fecha"].dt.strftime("%Y-%m")

for col, df, desc, fuente in series_list:
    n = df[df.columns[1]].notna().sum() if df is not None else 0
    print(f"  {col:30s}: {n:3d}/264 | {fuente}")
    if df is not None and len(df) > 0:
        val_col = df.columns[1]
        tmp = df[["fecha", val_col]].copy()
        tmp["fecha"] = tmp["fecha"].dt.to_period("M").dt.to_timestamp()
        maestro = maestro.merge(tmp.rename(columns={val_col: col}),
                                on="fecha", how="left")

print("=" * 62)

# Exportar
with pd.ExcelWriter(OUT_FILE, engine="openpyxl") as writer:
    maestro.to_excel(writer, sheet_name="series_mensuales", index=False)
    pd.DataFrame(
        [(col, desc, fuente, maestro[col].notna().sum() if col in maestro else 0,
          "Completa" if col in maestro and maestro[col].notna().sum() >= 260 else "Parcial")
         for col, df, desc, fuente in series_list],
        columns=["variable","descripcion","fuente","obs_disponibles","estado"]
    ).to_excel(writer, sheet_name="metadata", index=False)
    pd.DataFrame([
        {"Variable":"Credito (Nov-Dic 2025)",
         "Nota":"BCCh publica con lag ~2 meses. Datos disponibles hasta Oct 2025.",
         "Accion":"Volver a correr este script en Feb 2026 para obtener datos completos."},
    ]).to_excel(writer, sheet_name="notas", index=False)

print(f"\nArchivo actualizado: {OUT_FILE}")
