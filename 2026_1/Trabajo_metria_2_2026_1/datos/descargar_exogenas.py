"""
Descarga FEDFUNDS y CPIAUCSL desde la API pública de FRED.
Guarda ambas series en variables_exogenas.xlsx y genera gráficos de verificación.
"""

import requests
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

# ── Configuración ──────────────────────────────────────────────────────────────
OUTPUT_DIR   = Path(__file__).parent          # carpeta datos/
EXCEL_PATH   = OUTPUT_DIR / "variables_exogenas.xlsx"
PLOT_DIR     = OUTPUT_DIR
START_DATE   = "2004-01-01"
END_DATE     = datetime.today().strftime("%Y-%m-%d")

FRED_BASE = "https://fred.stlouisfed.org/graph/fredgraph.csv?id="

SERIES = {
    "FEDFUNDS": {
        "sheet":       "Fed_Funds_Rate",
        "label":       "Tasa de Fondos Federales (Fed Funds Rate)",
        "unidad":      "Porcentaje anual (%)",
        "fuente":      "FRED – Federal Reserve Bank of St. Louis",
        "color":       "#1F4E79",
        "ylabel":      "Tasa (%)",
        "descripcion": "Tasa de política monetaria de EE.UU. (nivel, promedio mensual)",
    },
    "CPIAUCSL": {
        "sheet":       "CPI_EEUU",
        "label":       "IPC Estados Unidos (CPI All Urban Consumers)",
        "unidad":      "Índice (base 1982–84 = 100)",
        "fuente":      "FRED – Federal Reserve Bank of St. Louis",
        "color":       "#375623",
        "ylabel":      "Índice (1982–84 = 100)",
        "descripcion": "IPC total EE.UU. en nivel (sin diferenciar ni transformar)",
    },
}

# ── Descarga ───────────────────────────────────────────────────────────────────
dataframes = {}

for series_id, meta in SERIES.items():
    print(f"\nDescargando {series_id}...")
    url = FRED_BASE + series_id
    r = requests.get(url, timeout=30)
    r.raise_for_status()

    from io import StringIO
    df = pd.read_csv(StringIO(r.text))
    df.columns = ["Fecha", "Valor"]
    df["Fecha"] = pd.to_datetime(df["Fecha"])
    df["Valor"] = pd.to_numeric(df["Valor"], errors="coerce")

    # Filtrar: desde 2004-01-01 hasta hoy; eliminar filas sin valor
    df = df[(df["Fecha"] >= START_DATE) & (df["Fecha"] <= END_DATE)]
    df = df.dropna(subset=["Valor"])
    df = df.sort_values("Fecha").reset_index(drop=True)

    dataframes[series_id] = (df, meta)
    print(f"  OK - {len(df)} obs.  "
          f"({df['Fecha'].min().strftime('%Y-%m')} a {df['Fecha'].max().strftime('%Y-%m')})")
    print(f"  Rango de valores: {df['Valor'].min():.2f} – {df['Valor'].max():.2f}")

# ── Excel ──────────────────────────────────────────────────────────────────────
print("\nGenerando Excel...")

with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", datetime_format="YYYY-MM-DD") as writer:
    for series_id, (df, meta) in dataframes.items():
        df.to_excel(writer, sheet_name=meta["sheet"], index=False)

# Formatear con openpyxl
wb = load_workbook(EXCEL_PATH)

HEADER_BLUE  = PatternFill("solid", fgColor="1F4E79")
HEADER_GREEN = PatternFill("solid", fgColor="375623")
ALT_FILL     = PatternFill("solid", fgColor="D9E1F2")
ALT_GREEN    = PatternFill("solid", fgColor="E2EFDA")
WHITE_FONT   = Font(bold=True, color="FFFFFF", size=11)
BORDER_THIN  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

for series_id, (df, meta) in dataframes.items():
    ws = wb[meta["sheet"]]

    # Encabezado de info (fila 1 antes de los datos)
    ws.insert_rows(1, amount=4)

    # Fila 1: título de la serie
    ws["A1"] = meta["label"]
    ws["A1"].font = Font(bold=True, size=13, color="1F4E79")
    ws.merge_cells("A1:B1")

    # Filas 2-4: metadatos
    meta_rows = [
        ("Unidad",      meta["unidad"]),
        ("Fuente",      meta["fuente"]),
        ("Periodo",     f"{df['Fecha'].min().strftime('%Y-%m')} a {df['Fecha'].max().strftime('%Y-%m')}  |  T = {len(df)} obs."),
    ]
    for i, (k, v) in enumerate(meta_rows, start=2):
        ws.cell(row=i, column=1, value=k).font  = Font(bold=True, size=10)
        ws.cell(row=i, column=2, value=v).font  = Font(size=10)

    # Fila 5 (encabezado de columnas, desplazado por insert_rows)
    header_row = 5
    hdr_fill = HEADER_BLUE if series_id == "FEDFUNDS" else HEADER_GREEN
    alt_fill  = ALT_FILL   if series_id == "FEDFUNDS" else ALT_GREEN

    for col in [1, 2]:
        cell = ws.cell(row=header_row, column=col)
        cell.fill   = hdr_fill
        cell.font   = WHITE_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER_THIN

    # Datos con filas alternadas
    for row_idx in range(header_row + 1, header_row + 1 + len(df)):
        for col in [1, 2]:
            cell = ws.cell(row=row_idx, column=col)
            cell.border = BORDER_THIN
            cell.alignment = Alignment(horizontal="center")
            if (row_idx - header_row) % 2 == 0:
                cell.fill = alt_fill
        # Formato fecha
        date_cell = ws.cell(row=row_idx, column=1)
        date_cell.number_format = "YYYY-MM-DD"
        # Formato numérico
        val_cell = ws.cell(row=row_idx, column=2)
        val_cell.number_format = "0.00"

    # Anchos de columna
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 18

    # Freeze
    ws.freeze_panes = f"A{header_row + 1}"

wb.save(EXCEL_PATH)
print(f"  Excel guardado: {EXCEL_PATH.name}")
print(f"  Hojas: {wb.sheetnames}")

# ── Gráficos ───────────────────────────────────────────────────────────────────
print("\nGenerando gráficos de verificación...")

fig, axes = plt.subplots(2, 1, figsize=(13, 9))
fig.suptitle(
    "Variables Exógenas – Verificación de Datos en Nivel\n(Fuente: FRED, Federal Reserve Bank of St. Louis)",
    fontsize=13, fontweight="bold", y=0.98
)

for ax, (series_id, (df, meta)) in zip(axes, dataframes.items()):
    color = meta["color"]

    ax.plot(df["Fecha"], df["Valor"],
            color=color, linewidth=1.6, alpha=0.92, zorder=3)
    ax.fill_between(df["Fecha"], df["Valor"],
                    alpha=0.12, color=color, zorder=2)

    # Línea vertical: inicio de la muestra del VAR
    ax.axvline(pd.Timestamp("2004-01-01"), color="gray",
               linestyle="--", linewidth=0.9, alpha=0.7,
               label="Inicio muestra VAR (2004m1)")

    # Marcar el último valor
    last_date = df["Fecha"].iloc[-1]
    last_val  = df["Valor"].iloc[-1]
    ax.annotate(
        f"Último: {last_val:.2f}\n({last_date.strftime('%b %Y')})",
        xy=(last_date, last_val),
        xytext=(-70, 15), textcoords="offset points",
        fontsize=8.5, color=color, fontweight="bold",
        arrowprops=dict(arrowstyle="-", color=color, lw=0.8),
    )

    ax.set_title(meta["label"], fontsize=11, fontweight="bold", pad=6)
    ax.set_ylabel(meta["ylabel"], fontsize=9)
    ax.set_xlabel("")
    ax.xaxis.set_major_locator(mdates.YearLocator(2))
    ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
    ax.tick_params(axis="x", labelsize=8, rotation=30)
    ax.tick_params(axis="y", labelsize=8)
    ax.grid(axis="y", linestyle="--", alpha=0.4, linewidth=0.7)
    ax.spines[["top","right"]].set_visible(False)
    ax.legend(fontsize=8, loc="upper left")

    # Cuadro de estadísticas descriptivas
    stats_text = (
        f"Media: {df['Valor'].mean():.2f}  |  "
        f"Mín: {df['Valor'].min():.2f}  |  "
        f"Máx: {df['Valor'].max():.2f}  |  "
        f"T = {len(df)} obs."
    )
    ax.text(0.01, 0.03, stats_text,
            transform=ax.transAxes, fontsize=7.5,
            color="gray", verticalalignment="bottom",
            bbox=dict(boxstyle="round,pad=0.3", facecolor="white",
                      edgecolor="lightgray", alpha=0.8))

plt.tight_layout(rect=[0, 0, 1, 0.96])

plot_path = PLOT_DIR / "grafico_variables_exogenas.png"
plt.savefig(plot_path, dpi=150, bbox_inches="tight")
plt.close()
print(f"  Gráfico guardado: {plot_path.name}")

# ── Resumen final ──────────────────────────────────────────────────────────────
print("\n" + "=" * 60)
print("RESUMEN")
print("=" * 60)
for series_id, (df, meta) in dataframes.items():
    print(f"\n  {series_id}  –  {meta['unidad']}")
    print(f"    Observaciones : {len(df)}")
    print(f"    Periodo       : {df['Fecha'].min().strftime('%Y-%m')} a {df['Fecha'].max().strftime('%Y-%m')}")
    print(f"    Media         : {df['Valor'].mean():.4f}")
    print(f"    Mínimo        : {df['Valor'].min():.4f}")
    print(f"    Máximo        : {df['Valor'].max():.4f}")
    print(f"    Último valor  : {df['Valor'].iloc[-1]:.4f}  ({df['Fecha'].iloc[-1].strftime('%Y-%m')})")

print(f"\nArchivos generados en: {OUTPUT_DIR}")
print(f"  {EXCEL_PATH.name}")
print(f"  {plot_path.name}")
print("=" * 60)
